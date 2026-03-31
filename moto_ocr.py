import requests
import re
import os
import base64
import openpyxl
import shutil
import datetime
from dotenv import load_dotenv

load_dotenv()

# 百度OCR配置
API_KEY = os.getenv("BAIDU_API_KEY")
SECRET_KEY = os.getenv("BAIDU_SECRET_KEY")

if not API_KEY or not SECRET_KEY:
    raise ValueError("请在 .env 文件中配置 BAIDU_API_KEY 和 BAIDU_SECRET_KEY")

# OCR接口URL列表，按优先级排序
OCR_URLS = [
    "https://aip.baidubce.com/rest/2.0/ocr/v1/accurate_basic",  # 高精度版
    "https://aip.baidubce.com/rest/2.0/ocr/v1/general_basic",   # 标准版
    "https://aip.baidubce.com/rest/2.0/ocr/v1/accurate",       # 高精度含位置版
    "https://aip.baidubce.com/rest/2.0/ocr/v1/general"         # 标准含位置版
]

def get_access_token():
    url = "https://aip.baidubce.com/oauth/2.0/token"
    params = {
        "grant_type": "client_credentials",
        "client_id": API_KEY,
        "client_secret": SECRET_KEY
    }
    response = requests.post(url, params=params)
    if response.status_code == 200:
        return response.json()["access_token"]
    else:
        return print("获取access_token失败")

def ocr_image(image_path, token):
    with open(image_path, "rb") as f:
        img = base64.b64encode(f.read()).decode()
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    payload = {"image": img, "access_token": token}
    
    for ocr_url in OCR_URLS:
        print(f"当前接口：${ocr_url}")
        try:
            response = requests.post(ocr_url, headers=headers, data=payload)
            if response.status_code != 200:
                print(f"接口请求失败，状态码：{response.status_code}")
                continue
                
            result = response.json()
            if 'error_code' in result:
                if result['error_code'] in [17, 18]:
                    print(f"接口配额不足，尝试下一个接口\n")
                    continue
                print(f"接口返回错误：{result['error_msg']}")
                continue
            
            if 'words_result' in result:
                if isinstance(result['words_result'], list):
                    return "\n".join([item["words"] for item in result["words_result"]])
                return "\n".join([item["words"] for item in result["words_result"].values()])
        except Exception as e:
            print(f"调用接口时发生错误：{str(e)}")
    
    return None

def extract_info(text):
    # 处理带冒号和不带冒号的姓名格式
    name_patterns = [
        re.compile(r"姓名：(\S+)"),
        re.compile(r"姓名\s*[\n\r]*(\S+)")
    ]
    
    name = "未识别到姓名"
    for pattern in name_patterns:
        match = pattern.search(text)
        if match:
            name = match.group(1)
            break
    
    # 处理带冒号和不带冒号的金额格式
    amount_patterns = [
        re.compile(r"实付金额：￥(\d+(?:\.\d{2})?)"),
        re.compile(r"实付金额\s*[\n\r]*￥(\d+(?:\.\d{2})?)"),
    ]
    
    amount = "未识别到金额"
    for pattern in amount_patterns:
        match = pattern.search(text)
        if match:
            # 如果金额包含小数点，去掉小数部分
            amount = match.group(1).split('.')[0]
            break

    return {
        "姓名": name,
        "实付金额": amount,
    }

def move_to_processed(image_path, image_folder):
    # 获取相对路径
    rel_path = os.path.relpath(image_path, image_folder)
    path_parts = rel_path.split(os.sep)
    
    # 获取当前日期作为一级目录
    current_date = datetime.datetime.now().strftime("%m月%d日")
    
    # 从原始路径中提取人名和群名
    # 假设路径结构为: 图片/人名/群名/图片文件
    if len(path_parts) >= 3:
        person_name = path_parts[0]  # 人名目录
        group_name = path_parts[1]   # 群名目录
        image_name = path_parts[-1]  # 图片文件名
    else:
        person_name = "未分类"
        group_name = "默认群组"
        image_name = path_parts[-1]

    # 从原始路径中获取正确的人名
    original_person_name = os.path.basename(os.path.dirname(os.path.dirname(image_path)))
    
    # 创建并获取归档目录路径
    archive_base = os.path.join(os.path.dirname(image_folder), "图片归档")
    if not os.path.exists(archive_base):
        os.makedirs(archive_base)
    
    # 构建新的归档路径：图片归档/日期/人名/群名/图片
    archive_path = os.path.join(archive_base, current_date, original_person_name, group_name, image_name)
    
    # 确保归档目录存在
    os.makedirs(os.path.dirname(archive_path), exist_ok=True)
    
    # 移动文件到归档目录
    shutil.move(image_path, archive_path)
    
    return archive_path


def get_unique_filename(file_path):
    # 如果文件不存在，直接返回原始路径
    if not os.path.exists(file_path):
        return file_path
    
    # 分离文件名和扩展名
    directory = os.path.dirname(file_path)
    filename = os.path.basename(file_path)
    name, ext = os.path.splitext(filename)
    
    # 尝试添加数字后缀直到找到一个不存在的文件名
    counter = 1
    while True:
        new_filename = os.path.join(directory, f"{name}{counter}{ext}")
        if not os.path.exists(new_filename):
            return new_filename
        counter += 1

def batch_process(image_folder, output_file = "result.xlsx"):
    access_token = get_access_token()
    # 使用字典来存储不同群组的结果
    group_results = {}
    
    for root, dirs, files in os.walk(image_folder):
        for filename in files:
            if filename.lower().endswith((".png", ".jpg", ".jpeg")):
                image_path = os.path.join(root, filename)
                text = ocr_image(image_path, access_token)
                info = extract_info(text)
                
                # 移动文件到processed目录
                processed_path = move_to_processed(image_path, image_folder)
                
                # 使用处理后的路径更新信息
                info["图片名称"] = os.path.relpath(processed_path, os.path.dirname(image_folder))
                folder_path = os.path.dirname(os.path.relpath(processed_path, os.path.dirname(image_folder)))
                info["所属文件夹"] = folder_path if folder_path else "根目录"
                
                # 获取群组信息
                path_parts = folder_path.split(os.sep)
                if len(path_parts) > 2:  # 如果有多于2层目录，最后一层是群名
                    group_name = path_parts[-1]
                else:
                    group_name = "默认群组"
                
                # 获取日期文件夹路径和人名文件夹路径
                date_folder = os.path.join(os.path.dirname(image_folder), "结果", 
                                          datetime.datetime.now().strftime("%m月%d日"))
                # 从原始路径中获取正确的人名
                original_person_name = os.path.basename(os.path.dirname(os.path.dirname(image_path)))
                result_folder = os.path.join(date_folder, original_person_name)  # 将人名添加到结果路径中
                
                # 将结果添加到对应群组
                if group_name not in group_results:
                    group_results[group_name] = {
                        "results": [],
                        "date_folder": result_folder  # 更新为包含人名的路径
                    }
                group_results[group_name]["results"].append(info)
                
                import time
                time.sleep(0.5)
    
    # 为每个群组创建单独的Excel文件
    headers = ["序号", "客户姓名", "交强险", "意外险", "总金额", "佣金", "日期", "出单点"]
    
    # 用于统计佣金的字典
    commission_stats = {}
    date_folders = set()
    
    # 获取当前日期
    current_date = datetime.datetime.now()
    # 构建月度台账文件名
    ledger_filename = f"{current_date.year}年{current_date.month}月摩托车台账.xlsx"
    ledger_path = os.path.join(os.path.dirname(os.path.abspath("./图片")), "结果", ledger_filename)
    
    # 如果月度台账不存在，复制模板文件
    template_path = os.path.join(os.path.dirname(os.path.abspath("./图片")), "结果", "2025年3月摩托车台账.xlsx")
    if not os.path.exists(ledger_path) and os.path.exists(template_path):
        shutil.copy(template_path, ledger_path)
    
    # 打开月度台账
    try:
        ledger_wb = openpyxl.load_workbook(ledger_path)
        ledger_ws = ledger_wb.active
    except Exception as e:
        print(f"无法打开月度台账：{str(e)}")
        return
    
    # 获取月度台账中最后一个有效行的序号
    last_row = 2  # 从第二行开始查找（跳过表头）
    last_number = 0
    for row in range(2, ledger_ws.max_row + 1):  # 从第二行开始遍历
        cell_value = ledger_ws.cell(row=row, column=1).value
        if isinstance(cell_value, (int, float)) or (isinstance(cell_value, str) and cell_value.isdigit()):
            last_row = row
            last_number = int(float(cell_value))
    
    for group_name, data in group_results.items():
        # 确保日期文件夹存在
        os.makedirs(data["date_folder"], exist_ok=True)
        date_folders.add(data["date_folder"])
        
        # 创建Excel文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        
        # 设置列宽
        ws.column_dimensions['A'].width = 10  # 序号列
        ws.column_dimensions['B'].width = 20  # 客户姓名列
        ws.column_dimensions['C'].width = 15  # 交强险列
        ws.column_dimensions['D'].width = 15  # 意外险列
        ws.column_dimensions['E'].width = 15  # 总金额列
        ws.column_dimensions['F'].width = 15  # 佣金列
        ws.column_dimensions['G'].width = 15  # 日期列
        ws.column_dimensions['H'].width = 20  # 出单点列
        
        # 统计变量
        count_60 = 0
        count_20 = 0
        total_commission = 0
        
        # 添加数据
        current_date_str = current_date.strftime("%Y.%m.%d")
        for idx, result in enumerate(data["results"], 1):
            total_amount = int(result["实付金额"]) if result["实付金额"] != "未识别到金额" else 0
            
            # 获取当前月份的日期
            current_month = current_date.month
            tax_amount = 3 * (13 - current_month) if current_month <= 13 else 0
            
            # 定义七种金额情况
            cases = [
                {"total": 392, "compulsory": 156, "accident": 200, "tax": 0},
                {"total": 356, "compulsory": 156, "accident": 200, "tax": tax_amount},
                {"total": 256, "compulsory": 156, "accident": 100, "tax": tax_amount},
                {"total": 204, "compulsory": 104, "accident": 100, "tax": 0},
                {"total": 304, "compulsory": 104, "accident": 200, "tax": 0},
                {"total": 356, "compulsory": 156, "accident": 200, "tax": 0},
                {"total": 256, "compulsory": 156, "accident": 100, "tax": 0}
            ]
            
            # 匹配金额情况
            matched = False
            compulsory_insurance = 0
            accident_insurance = 0
            
            for case in cases:
                # 计算实际需要比较的总金额（包含税额）
                actual_total = case["total"] + case["tax"]
                if total_amount == actual_total:
                    compulsory_insurance = case["compulsory"]
                    accident_insurance = case["accident"]
                    matched = True
                    break
            
            # 如果没有匹配到任何情况，输出错误信息
            if not matched:
                print(f"【{original_person_name}】的【{group_name}】中出现金额不匹配保单：【{result['姓名']}】，金额【{total_amount}】")
            
            # 计算佣金
            commission = 0
            if accident_insurance == 100:
                commission = 5
                count_20 += 1
            elif accident_insurance == 200:
                commission = 30
                count_60 += 1
            
            total_commission += commission
            
            row_data = [
                idx,
                result["姓名"],
                compulsory_insurance,
                accident_insurance,
                total_amount,
                commission,
                current_date_str,
                group_name
            ]
            ws.append(row_data)
            
            # 同步数据到月度台账
            last_number += 1
            ledger_data = [
                last_number,  # 序号
                current_date_str,  # 出单日期
                result["姓名"],  # 被保险人
                compulsory_insurance,  # 交强险保费
                accident_insurance,  # 意外险保费
                total_amount,  # 保费合计
                group_name,  # 出单点
                ""  # 备注
            ]
            last_row += 1
            # 写入数据到月度台账，处理合并单元格
            # 创建字体和边框样式
            font = openpyxl.styles.Font(name='等线', size=20, bold=True)
            border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            for col, value in enumerate(ledger_data, 1):
                cell = ledger_ws.cell(row=last_row, column=col)
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    # 获取合并区域的主单元格
                    for merged_range in ledger_ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            main_cell = ledger_ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            main_cell.value = value
                            main_cell.font = font
                            main_cell.border = border
                            break
                else:
                    cell.value = value
                    cell.font = font
                    cell.border = border
        
        # 保存文件
        # 生成不重复的Excel文件名
        excel_path = os.path.join(data["date_folder"], f"{group_name}.xlsx")
        excel_path = get_unique_filename(excel_path)
        wb.save(excel_path)
        
        # 保存统计信息
        commission_stats[group_name] = {
            "count_60": count_60,
            "count_20": count_20,
            "total_commission": total_commission,
            "date_folder": data["date_folder"]
        }
    
    # 保存月度台账
    try:
        ledger_wb.save(ledger_path)
        ledger_wb.close()
        print(f"\n月度台账已更新：{ledger_path}")
    except Exception as e:
        print(f"保存月度台账时出错：{str(e)}")
    
    # 生成结算统计文件
    # 按人名和日期组织数据
    person_date_stats = {}
    for group_name, stats in commission_stats.items():
        date_folder = stats["date_folder"]
        person_name = os.path.basename(date_folder)  # 从路径中获取人名
        date_path = os.path.dirname(date_folder)    # 获取日期路径
        
        if date_path not in person_date_stats:
            person_date_stats[date_path] = {}
        
        if person_name not in person_date_stats[date_path]:
            person_date_stats[date_path][person_name] = {
                "groups": [],
                "total_count_60": 0,
                "total_count_20": 0,
                "total_commission": 0,
                "folder_path": date_folder
            }
        
        if stats["count_60"] > 0 or stats["count_20"] > 0:
            group_detail = []
            commission_parts = []
            if stats['count_60'] > 0:
                group_detail.append(f"{stats['count_60']}单30")
                commission_parts.append(f"{stats['count_60']}×30")
            if stats['count_20'] > 0:
                group_detail.append(f"{stats['count_20']}单5元")
                commission_parts.append(f"{stats['count_20']}×5")
            
            # 构建群组文本
            group_text = f"{group_name} {', '.join(group_detail)}"
            if len(commission_parts) > 1:
                group_text += f"，{' + '.join(commission_parts)} = {stats['total_commission']}"
            group_text += f"，付{stats['total_commission']}元"
            
            person_stats = person_date_stats[date_path][person_name]
            person_stats["groups"].append(group_text)
            person_stats["total_count_60"] += stats["count_60"]
            person_stats["total_count_20"] += stats["count_20"]
            person_stats["total_commission"] += stats["total_commission"]
    
    # 为每个人生成结算文件
    for date_path, person_stats in person_date_stats.items():
        for person_name, stats in person_stats.items():
            if stats["total_count_60"] > 0 or stats["total_count_20"] > 0:
                settlement_content = []
                
                # 添加每个群的统计信息
                settlement_content.extend(stats["groups"])
                
                # 添加总计信息
                total_parts = []
                commission_parts = []
                if stats["total_count_60"] > 0:
                    total_parts.append(f"{stats['total_count_60']}单30")
                    commission_parts.append(f"{stats['total_count_60']}×30")
                if stats["total_count_20"] > 0:
                    total_parts.append(f"{stats['total_count_20']}单5元")
                    commission_parts.append(f"{stats['total_count_20']}×5")
                
                total_text = f"总共 {', '.join(total_parts)}"
                if len(commission_parts) > 1:
                    total_text += f"，{' + '.join(commission_parts)} = {stats['total_commission']}"
                total_text += f"，付{stats['total_commission']}元"
                settlement_content.append(total_text)
                
                # 写入结算文件
                # 生成不重复的结算文件名
                settlement_file = os.path.join(stats["folder_path"], "结算.txt")
                settlement_file = get_unique_filename(settlement_file)
                with open(settlement_file, "w", encoding="utf-8") as f:
                    f.write("\n".join(settlement_content))

batch_process("./图片")
