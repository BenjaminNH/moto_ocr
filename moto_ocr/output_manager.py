import os
import shutil
import datetime
import openpyxl
from moto_ocr.commission import match_premium, calculate_commission
from moto_ocr.config import EXCEL_HEADERS, COLUMN_WIDTHS, LEDGER_TEMPLATE_NAME
from moto_ocr.file_utils import get_unique_filename
from moto_ocr.settlement import build_group_text, build_total_text

_LEDGER_FONT = openpyxl.styles.Font(name='等线', size=20, bold=True)
_LEDGER_BORDER = openpyxl.styles.Border(
    left=openpyxl.styles.Side(style='thin'),
    right=openpyxl.styles.Side(style='thin'),
    top=openpyxl.styles.Side(style='thin'),
    bottom=openpyxl.styles.Side(style='thin')
)


def move_to_processed(image_path, image_folder):
    """移动图片到归档目录。
    
    Args:
        image_path: 图片原始路径
        image_folder: 图片根目录（如 "./图片"）
        
    Returns:
        str: 归档后的路径
    """
    rel_path = os.path.relpath(image_path, image_folder)
    path_parts = rel_path.split(os.sep)
    current_date = datetime.datetime.now().strftime("%m月%d日")
    
    if len(path_parts) >= 3:
        group_name = path_parts[1]
        image_name = path_parts[-1]
    else:
        group_name = "默认群组"
        image_name = path_parts[-1]

    original_person_name = os.path.basename(os.path.dirname(os.path.dirname(image_path)))
    
    archive_base = os.path.join(os.path.dirname(image_folder), "图片归档")
    archive_path = os.path.join(archive_base, current_date, original_person_name, group_name, image_name)
    
    os.makedirs(os.path.dirname(archive_path), exist_ok=True)
    shutil.move(image_path, archive_path)
    
    return archive_path


def create_ledger(ledger_path, template_path):
    """创建或打开月度台账。
    
    Args:
        ledger_path: 月度台账目标路径
        template_path: 模板文件路径
        
    Returns:
        tuple: (workbook, worksheet) 或 (None, None)
    """
    if not os.path.exists(ledger_path) and os.path.exists(template_path):
        shutil.copy(template_path, ledger_path)
    
    try:
        wb = openpyxl.load_workbook(ledger_path)
        ws = wb.active
        return wb, ws
    except Exception as e:
        print(f"无法打开月度台账：{str(e)}")
        return None, None


def find_last_ledger_row(ws):
    """查找台账中最后一个有效行的序号。
    
    Returns:
        tuple: (last_row_index, last_number)
    """
    last_row = 2
    last_number = 0
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if isinstance(cell_value, (int, float)) or (isinstance(cell_value, str) and cell_value.isdigit()):
            last_row = row
            last_number = int(float(cell_value))
    return last_row, last_number


def write_to_ledger(ws, last_row, ledger_data):
    """写入一行数据到台账，处理合并单元格。
    
    Returns:
        int: 新的行号
    """
    last_row += 1
    for col, value in enumerate(ledger_data, 1):
        cell = ws.cell(row=last_row, column=col)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    main_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    main_cell.value = value
                    main_cell.font = _LEDGER_FONT
                    main_cell.border = _LEDGER_BORDER
                    break
        else:
            cell.value = value
            cell.font = _LEDGER_FONT
            cell.border = _LEDGER_BORDER
    
    return last_row


def create_group_excel(group_name, results, date_folder, current_date_str, month):
    """为单个群组创建Excel文件。
    
    Returns:
        tuple: (excel_path, commission_stats)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(EXCEL_HEADERS)
    
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width
    
    count_30 = 0
    count_50 = 0
    total_commission = 0
    
    for idx, result in enumerate(results, 1):
        raw_amount = result["实付金额"]
        total_amount = int(raw_amount) if raw_amount != "未识别到金额" else 0
        
        matched = match_premium(total_amount, month=month)
        if matched:
            compulsory_insurance = matched["compulsory"]
            accident_insurance = matched["accident"]
        else:
            compulsory_insurance = 0
            accident_insurance = 0
        
        commission = calculate_commission(accident_insurance)
        if accident_insurance == 200:
            count_30 += 1
        elif accident_insurance == 300:
            count_50 += 1
        total_commission += commission
        
        row_data = [idx, result["姓名"], compulsory_insurance, accident_insurance,
                    total_amount, commission, current_date_str, group_name]
        ws.append(row_data)
    
    excel_path = os.path.join(date_folder, f"{group_name}.xlsx")
    excel_path = get_unique_filename(excel_path)
    wb.save(excel_path)
    
    stats = {
        "count_30": count_30,
        "count_50": count_50,
        "total_commission": total_commission,
        "date_folder": date_folder,
    }
    
    return excel_path, stats


def write_ledger_row_for_result(ws, last_row, last_number, result, compulsory_insurance,
                                accident_insurance, total_amount, group_name, current_date_str):
    """为单个结果写入台账行。
    
    Returns:
        int: 新的 last_number
    """
    last_number += 1
    ledger_data = [
        last_number, current_date_str, result["姓名"],
        compulsory_insurance, accident_insurance, total_amount,
        group_name, ""
    ]
    new_row = write_to_ledger(ws, last_row, ledger_data)
    return last_number, new_row


def generate_settlement_files(commission_stats):
    """根据佣金统计生成结算文件。
    
    Args:
        commission_stats: 群组佣金统计字典
    """
    person_date_stats = {}
    for group_name, stats in commission_stats.items():
        date_folder = stats["date_folder"]
        person_name = os.path.basename(date_folder)
        date_path = os.path.dirname(date_folder)
        
        if date_path not in person_date_stats:
            person_date_stats[date_path] = {}
        
        if person_name not in person_date_stats[date_path]:
            person_date_stats[date_path][person_name] = {
                "groups": [],
                "total_count_30": 0,
                "total_count_50": 0,
                "total_commission": 0,
                "folder_path": date_folder
            }
        
        if stats["count_30"] > 0 or stats["count_50"] > 0:
            group_text = build_group_text(
                group_name, stats["count_30"], stats["count_50"], stats["total_commission"]
            )
            person_stats = person_date_stats[date_path][person_name]
            person_stats["groups"].append(group_text)
            person_stats["total_count_30"] += stats["count_30"]
            person_stats["total_count_50"] += stats["count_50"]
            person_stats["total_commission"] += stats["total_commission"]
    
    for date_path, person_stats in person_date_stats.items():
        for person_name, stats in person_stats.items():
            if stats["total_count_30"] > 0 or stats["total_count_50"] > 0:
                settlement_content = list(stats["groups"])
                total_text = build_total_text(
                    stats["total_count_30"], stats["total_count_50"], stats["total_commission"]
                )
                settlement_content.append(total_text)
                
                settlement_file = os.path.join(stats["folder_path"], "结算.txt")
                settlement_file = get_unique_filename(settlement_file)
                try:
                    with open(settlement_file, "w", encoding="utf-8") as f:
                        f.write("\n".join(settlement_content))
                except Exception as e:
                    print(f"保存结算文件失败：{settlement_file}，错误：{e}")
